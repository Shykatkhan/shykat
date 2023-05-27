import requests, json
import openpyxl
from datetime import datetime


def google_search(key):
    URL="http://suggestqueries.google.com/complete/search?client=firefox&q=" +key
    headers = {'User-agent':'Mozilla/5.0'} 
    response = requests.get(URL,headers=headers,verify='false') 
    result = json.loads(response.content.decode('utf-8'))
    return(result[1])


dt = datetime.now()
print('start operation......')
   
   
dataframe = openpyxl.load_workbook("Excel.xlsx")
  

sheet = dataframe['{}'.format(str(dt.strftime('%A')))]
for row in range(3, sheet.max_row+1):
    cell = sheet.cell(row=row, column=3)
    search_sugg=google_search(cell.value)
    cols4=sheet.cell(row=row, column=4)
    cols5=sheet.cell(row=row, column=5)
    cols4.value=str(max(search_sugg,key=len))
    cols5.value=str(min(search_sugg,key=len))
    
dataframe.save("Excel.xlsx")
print('operation done!')



# get current datetime